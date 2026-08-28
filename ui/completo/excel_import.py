"""
excel_import.py
----------------
Importação de pilar/cargas e perfil geotécnico a partir de uma planilha
Excel (`.xlsx`), layout fixo, documentado abaixo.

Restrição de a3-interface.md: esta tela não calcula nada. Cada função só lê
célula, converte tipo (texto -> float/str/enum) e chama os construtores de
`Pilar`/`CasoCarga`/`Esforcos`/`Camada`/`PerfilGeotecnico` já existentes em
`calc_core.sapata_isolada`. NENHUMA fórmula, correlação ou critério de
engenharia é reimplementado aqui.

ATENÇÃO (corrigido depois do defeito D4 do GATE 2, rodada 1): ao contrário
do que uma versão anterior deste docstring afirmava, NEM TODO construtor de
`calc_core.sapata_isolada` valida faixa/domínio sozinho. Só `Pilar`,
`Concreto` e `Aco` têm `__post_init__`; `Camada`, `PerfilGeotecnico`, `Solo`,
`Esforcos` e `CasoCarga` NÃO validam nada — aceitam, por exemplo, espessura
negativa ou N_SPT negativo sem erro. Como este módulo é a FRONTEIRA de dados
não confiáveis (planilha de terceiros), ele mesmo valida DOMÍNIO (não
engenharia: só que o número esteja no intervalo em que faz sentido existir)
antes de repassar ao construtor: espessura de camada tem de ser > 0;
nspt/Cc/e0/cv/Es e nível d'água têm de ser >= 0 quando preenchidos. Isso não
fere a regra "ui/ não calcula" — é validação de entrada, não cálculo. A
validação de domínio dentro do próprio `calc_core` (`__post_init__` em
`Camada`/`Solo`) é tarefa de outra rodada, de a4/a2.

Layout — aba "Pilar e cargas"
------------------------------
Cabeçalho na linha 1, dados a partir da linha 2, UMA LINHA POR CASO DE CARGA:

    ap (m) | bp (m) | caso | N (kN) | Mx (kN·m) | My (kN·m)

`ap`/`bp` só são lidos da PRIMEIRA linha de dados (linha 2) — podem ficar em
branco nas linhas seguintes. `caso` é o nome/rótulo do caso de carga; esta
função (`importar_pilar_e_cargas`) aceita QUALQUER texto aqui e devolve um
`CasoCarga` para cada linha, sem validar o nome. Só use um nome fora de "G"/
"Q"/"W" (maiúsculas, sem espaço nas bordas) se o destino for outro consumidor
de `list[CasoCarga]` — se o destino for `ui/completo/formulario.py::
PainelEntrada.preencher_casos` (o caso de `ui/completo/app.py::
_importar_excel`, que é o fluxo normal desta planilha), o nome PRECISA ser
exatamente "G", "Q" ou "W": a tela só tem esses três campos fixos, e
`preencher_casos` levanta `ValueError` — recusando a importação inteira, sem
preencher nada pela metade — para qualquer outro nome, em vez de descartá-lo
em silêncio. `N` é obrigatório em toda linha; `Mx`/`My` em branco valem 0.
Este layout NÃO tem colunas de Hx/Hy nem de tipo de ação (psi): cada linha
importada vira um `CasoCarga` com ação horizontal nula e o tipo padrão
(permanente) — ajuste tipo/psi/forças horizontais na tela depois de
importar, se o projeto precisar.

Layout — aba "Perfil geotécnico"
---------------------------------
Cabeçalho na linha 1, dados a partir da linha 2, UMA LINHA POR CAMADA, na
ORDEM DO TOPO PARA A BASE:

    camada | tipo | espessura (m) | nspt | Cc | e0 | cv (m²/ano) | Es (kPa) | nível d'água (m)

`tipo` precisa casar com um valor de `TipoSubstrato` ("granular", "coesivo",
"rocha", "aterro"); qualquer outro valor é rejeitado citando a linha e o
valor recebido. `Cc`/`e0`/`cv`/`Es` são os parâmetros de adensamento/
deformabilidade — só fazem sentido em camada coesiva (Cc/e0/cv) ou quando
não se quer depender da correlação com o SPT (Es); podem ficar em branco
nas demais linhas, e ficam em branco = lidos como `None`, nunca como zero.
`nível d'água` só é lida da PRIMEIRA linha de dados (mesma convenção de
`ap`/`bp` acima).

NOTA DE UNIDADE (cv): a coluna usa m²/ano, a MESMA unidade que
`calc_core.sapata_isolada.geotecnia.Camada.cv` espera (ver a docstring de
`Camada` e `recalques.tempo_para_grau`) — não cm²/s, unidade comum em outras
referências de geotecnia. Esta aba não faz NENHUMA conversão de unidade
(regra de a3-interface.md: "ui/ não calcula"): o número digitado na célula é
o número que entra direto no núcleo. Rotular a coluna com uma unidade
diferente da que o núcleo consome exigiria converter aqui, na UI — proibido
— ou entregaria ao núcleo um valor num fator de escala errado sem aviso,
que é o efeito colateral mais perigoso dos dois.

Use `gerar_modelo_importacao` para gerar um `.xlsx` de exemplo com as duas
abas de dados preenchidas e uma terceira aba de instruções.
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Font

from calc_core.sapata_isolada.acoes import CasoCarga, Esforcos, Pilar
from calc_core.sapata_isolada.geotecnia import Camada, PerfilGeotecnico, TipoSubstrato

ABA_PILAR = "Pilar e cargas"
ABA_PERFIL = "Perfil geotécnico"
ABA_INSTRUCOES = "Instruções"

CABECALHO_PILAR = ("ap (m)", "bp (m)", "caso", "N (kN)", "Mx (kN·m)", "My (kN·m)")
CABECALHO_PERFIL = ("camada", "tipo", "espessura (m)", "nspt", "Cc", "e0",
                    "cv (m²/ano)", "Es (kPa)", "nível d'água (m)")

TIPOS_ACEITOS = [t.value for t in TipoSubstrato]


class AbaAusente(ValueError):
    """A aba nomeada não existe na planilha aberta.

    Subclasse de `ValueError` de propósito: quem só quer tratar erro de
    importação em geral continua funcionando com `except ValueError`; quem
    precisa distinguir "aba ausente" (não-fatal para o perfil geotécnico,
    ver `ui/completo/app.py::_importar_excel`) usa este tipo específico.
    """


# --------------------------------------------------------------------------- #
#  Leitura de célula
# --------------------------------------------------------------------------- #
def _abrir(caminho) -> tuple[openpyxl.Workbook, openpyxl.Workbook]:
    """Abre a planilha DUAS VEZES: uma com `data_only=True` (valor
    CALCULADO de cada célula — o que toda leitura normal usa) e outra com
    `data_only=False` (a FÓRMULA em texto, sem avaliar).

    O segundo livro só serve para diagnosticar um caso que `data_only=True`
    sozinho não resolve: uma célula com FÓRMULA cujo valor calculado não
    está em cache (planilha montada por script/openpyxl, ou salva sem
    recalcular) faz `data_only=True` devolver `None` — indistinguível de
    "célula realmente vazia" sem olhar o livro bruto. Ver `_celula_float`/
    `_celula_texto` e o defeito D9 do GATE 2, rodada 1.
    """
    def _abrir_um(data_only: bool) -> openpyxl.Workbook:
        try:
            return openpyxl.load_workbook(caminho, data_only=data_only)
        except FileNotFoundError as erro:
            raise ValueError(
                f"Arquivo Excel não encontrado: {caminho!r}.") from erro
        except Exception as erro:
            raise ValueError(
                f"Não foi possível abrir {caminho!r} como planilha Excel "
                f"(.xlsx): {erro}") from erro

    return _abrir_um(True), _abrir_um(False)


def _aba(livro, nome: str):
    if nome not in livro.sheetnames:
        raise AbaAusente(
            f"A planilha não tem a aba {nome!r}. Abas encontradas: "
            f"{', '.join(livro.sheetnames)}.")
    return livro[nome]


def _vazia(valor) -> bool:
    return valor is None or (isinstance(valor, str) and not valor.strip())


def _formula_sem_cache(ws_bruto, linha: int, coluna: int) -> str | None:
    """Devolve o texto da fórmula (`"=..."`) se a célula BRUTA (sem avaliar)
    tiver uma fórmula — o sintoma de uma célula que parece vazia em
    `data_only=True` só porque o cache de valor calculado não existe.
    Devolve `None` se a célula bruta não for uma fórmula (está mesmo
    vazia, ou tem um valor literal)."""
    valor_bruto = ws_bruto.cell(row=linha, column=coluna).value
    if isinstance(valor_bruto, str) and valor_bruto.startswith("="):
        return valor_bruto
    return None


def _celula_float(ws, ws_bruto, linha: int, coluna: int, rotulo: str,
                  obrigatorio: bool, minimo: float | None = None,
                  estrito: bool = False) -> float | None:
    """`minimo`/`estrito` aplicam a validação de DOMÍNIO na fronteira
    (defeito D4 do GATE 2, rodada 1): `estrito=True` exige `valor > minimo`
    (ex.: espessura > 0); `estrito=False` exige `valor >= minimo` (ex.:
    nspt, Cc, e0, cv, Es, nível d'água >= 0). Ver docstring do módulo."""
    valor = ws.cell(row=linha, column=coluna).value
    if _vazia(valor):
        formula = _formula_sem_cache(ws_bruto, linha, coluna)
        if formula is not None:
            raise ValueError(
                f"Aba {ws.title!r}, linha {linha}, coluna {rotulo!r}: a "
                f"célula contém a fórmula {formula!r} sem valor calculado "
                "(cache vazio) — abra e salve a planilha no Excel/"
                "LibreOffice antes de importar, ou digite o valor "
                "diretamente na célula.")
        if obrigatorio:
            raise ValueError(
                f"Aba {ws.title!r}, linha {linha}, coluna {rotulo!r}: célula "
                "obrigatória vazia.")
        return None
    if isinstance(valor, bool):
        # ValueError de propósito, não TypeError (ruff: TRY004) — o
        # contrato público desta função (docstring de `importar_pilar_e_
        # cargas`/`importar_perfil_geotecnico` acima) é "sempre ValueError
        # para célula inválida", testado em
        # tests/test_projeto_e_excel.py::test_excel_valor_logico_bool_e_recusado
        # (ALTA #2 do GATE 2, rodada 2).
        raise ValueError(  # noqa: TRY004
            f"Aba {ws.title!r}, linha {linha}, coluna {rotulo!r}: valor "
            f"lógico ({valor!r}) onde se esperava número — confira se a "
            "célula não veio de uma fórmula lógica ou de um checkbox; "
            "digite o número diretamente.")
    if isinstance(valor, (int, float)):
        numero = float(valor)
    else:
        texto = str(valor).strip().replace(",", ".")
        try:
            numero = float(texto)
        except ValueError as erro:
            raise ValueError(
                f"Aba {ws.title!r}, linha {linha}, coluna {rotulo!r}: valor "
                f"{valor!r} não é numérico.") from erro
    if minimo is not None:
        valido = numero > minimo if estrito else numero >= minimo
        if not valido:
            relacao = "> " if estrito else ">= "
            raise ValueError(
                f"Aba {ws.title!r}, linha {linha}, coluna {rotulo!r}: valor "
                f"{numero!r} fora do domínio esperado (deve ser {relacao}"
                f"{minimo!r}).")
    return numero


def _celula_texto(ws, ws_bruto, linha: int, coluna: int, rotulo: str,
                  obrigatorio: bool = True) -> str | None:
    valor = ws.cell(row=linha, column=coluna).value
    if _vazia(valor):
        formula = _formula_sem_cache(ws_bruto, linha, coluna)
        if formula is not None:
            raise ValueError(
                f"Aba {ws.title!r}, linha {linha}, coluna {rotulo!r}: a "
                f"célula contém a fórmula {formula!r} sem valor calculado "
                "(cache vazio) — abra e salve a planilha no Excel/"
                "LibreOffice antes de importar, ou digite o valor "
                "diretamente na célula.")
        if obrigatorio:
            raise ValueError(
                f"Aba {ws.title!r}, linha {linha}, coluna {rotulo!r}: célula "
                "obrigatória vazia.")
        return None
    return str(valor).strip()


def _linha_em_branco(ws, linha: int, n_colunas: int) -> bool:
    return all(_vazia(ws.cell(row=linha, column=c).value)
              for c in range(1, n_colunas + 1))


# --------------------------------------------------------------------------- #
#  Importação
# --------------------------------------------------------------------------- #
def importar_pilar_e_cargas(caminho) -> tuple[Pilar, list[CasoCarga]]:
    """Lê a aba "Pilar e cargas" e devolve (`Pilar`, lista de `CasoCarga`).

    Levanta `ValueError` (ou `AbaAusente`, sua subclasse) se a aba não
    existir, se `ap`/`bp` faltarem na primeira linha de dados, se `N`
    faltar em alguma linha, se `caso` faltar, se algum valor numérico não
    puder ser convertido, ou se `ap`/`bp` aparecerem de novo, com valor
    DIFERENTE do da primeira linha, numa linha seguinte — sempre citando
    aba/linha/coluna.
    """
    livro, livro_bruto = _abrir(caminho)
    ws = _aba(livro, ABA_PILAR)
    ws_bruto = livro_bruto[ABA_PILAR]

    ap: float | None = None
    bp: float | None = None
    casos: list[CasoCarga] = []
    primeira = True
    for linha in range(2, ws.max_row + 1):
        if _linha_em_branco(ws, linha, 6):
            continue
        if primeira:
            ap = _celula_float(ws, ws_bruto, linha, 1, "ap (m)",
                               obrigatorio=True, minimo=0, estrito=True)
            bp = _celula_float(ws, ws_bruto, linha, 2, "bp (m)",
                               obrigatorio=True, minimo=0, estrito=True)
            primeira = False
        else:
            _exigir_valor_repetido_igual(ws, ws_bruto, linha, 1, "ap (m)", ap)
            _exigir_valor_repetido_igual(ws, ws_bruto, linha, 2, "bp (m)", bp)
        nome = _celula_texto(ws, ws_bruto, linha, 3, "caso", obrigatorio=True)
        N = _celula_float(ws, ws_bruto, linha, 4, "N (kN)", obrigatorio=True)
        Mx = _celula_float(ws, ws_bruto, linha, 5, "Mx (kN·m)",
                           obrigatorio=False) or 0.0
        My = _celula_float(ws, ws_bruto, linha, 6, "My (kN·m)",
                           obrigatorio=False) or 0.0
        casos.append(CasoCarga(nome=nome, esforcos=Esforcos(N=N, Mx=Mx, My=My)))

    if ap is None or bp is None or not casos:
        raise ValueError(
            f"Aba {ABA_PILAR!r}: nenhuma linha de dados encontrada (esperada "
            "ao menos uma linha a partir da linha 2, com ap, bp, caso e N "
            "preenchidos).")

    return Pilar(ap=ap, bp=bp), casos


def _exigir_valor_repetido_igual(ws, ws_bruto, linha: int, coluna: int,
                                 rotulo: str, valor_primeira_linha: float
                                 ) -> None:
    """`ap`/`bp` (aba "Pilar e cargas") e nível d'água (aba "Perfil
    geotécnico") só são LIDOS da primeira linha de dados — mas se uma
    linha seguinte trouxer um valor DIFERENTE, isso quase sempre é engano
    do usuário (ex.: pilares de projetos diferentes copiados na mesma
    planilha), não um valor a ignorar em silêncio. Célula vazia ou com o
    MESMO valor da primeira linha continua aceita sem aviso."""
    outro = _celula_float(ws, ws_bruto, linha, coluna, rotulo, obrigatorio=False)
    if outro is not None and outro != valor_primeira_linha:
        if valor_primeira_linha is None:
            # MEDIA #3 do GATE 2, rodada 2: mensagem específica para quando
            # a PRIMEIRA linha estava em branco (não "None" cru na
            # mensagem, que confundia mais do que ajudava).
            raise ValueError(
                f"Aba {ws.title!r}, linha {linha}, coluna {rotulo!r}: valor "
                f"{outro!r} conflita com a primeira linha de dados, que "
                "está em branco nesta coluna — esta coluna só é lida da "
                "PRIMEIRA linha; se o valor pertence ao projeto, preencha-o "
                "na PRIMEIRA linha de dados (não numa linha seguinte).")
        raise ValueError(
            f"Aba {ws.title!r}, linha {linha}, coluna {rotulo!r}: valor "
            f"{outro!r} conflita com o valor {valor_primeira_linha!r} já "
            "lido na primeira linha de dados — esta coluna só é lida da "
            "PRIMEIRA linha; deixe em branco nas linhas seguintes (ou "
            "repita o mesmo valor) em vez de preencher um valor diferente.")


def importar_perfil_geotecnico(caminho) -> PerfilGeotecnico:
    """Lê a aba "Perfil geotécnico" e devolve o `PerfilGeotecnico` completo,
    na mesma ordem das linhas da planilha (topo para a base).

    Levanta `AbaAusente` se a aba não existir, e `ValueError` para célula
    obrigatória vazia, valor numérico inválido, fora do domínio esperado
    (espessura <= 0; nspt/Cc/e0/cv/Es/nível d'água < 0), `tipo` fora do
    enum `TipoSubstrato`, ou nível d'água conflitante entre linhas —
    sempre citando aba/linha/coluna.
    """
    livro, livro_bruto = _abrir(caminho)
    ws = _aba(livro, ABA_PERFIL)
    ws_bruto = livro_bruto[ABA_PERFIL]

    nivel_agua: float | None = None
    primeira = True
    camadas: list[Camada] = []
    for linha in range(2, ws.max_row + 1):
        if _linha_em_branco(ws, linha, 9):
            continue
        nome = _celula_texto(ws, ws_bruto, linha, 1, "camada", obrigatorio=True)
        tipo_txt = _celula_texto(ws, ws_bruto, linha, 2, "tipo", obrigatorio=True)
        try:
            tipo = TipoSubstrato(tipo_txt)
        except ValueError as erro:
            raise ValueError(
                f"Aba {ABA_PERFIL!r}, linha {linha}, coluna 'tipo': valor "
                f"{tipo_txt!r} não é um tipo de substrato válido; use um de "
                f"{TIPOS_ACEITOS}.") from erro
        espessura = _celula_float(ws, ws_bruto, linha, 3, "espessura (m)",
                                  obrigatorio=True, minimo=0, estrito=True)
        nspt = _celula_float(ws, ws_bruto, linha, 4, "nspt",
                             obrigatorio=False, minimo=0)
        Cc = _celula_float(ws, ws_bruto, linha, 5, "Cc",
                           obrigatorio=False, minimo=0)
        e0 = _celula_float(ws, ws_bruto, linha, 6, "e0",
                           obrigatorio=False, minimo=0)
        cv = _celula_float(ws, ws_bruto, linha, 7, "cv (m²/ano)",
                           obrigatorio=False, minimo=0)
        Es = _celula_float(ws, ws_bruto, linha, 8, "Es (kPa)",
                           obrigatorio=False, minimo=0)
        if primeira:
            nivel_agua = _celula_float(ws, ws_bruto, linha, 9,
                                       "nível d'água (m)", obrigatorio=False,
                                       minimo=0)
            primeira = False
        else:
            # MEDIA #3 do GATE 2, rodada 2: a guarda antiga (`elif
            # nivel_agua is not None`) só checava conflito quando a
            # PRIMEIRA linha de dados já trazia um valor — com a linha 2
            # em branco e a linha 3 com 2,5, a checagem inteira era
            # PULADA (branch nunca entrava) e `importar_perfil_geotecnico`
            # devolvia `nivel_agua=None` em silêncio, subestimando o
            # recalque (lado INSEGURO; reproduzido no relatório com
            # 204,9mm em vez de 217,3mm). `_exigir_valor_repetido_igual`
            # já trata corretamente o caso `valor_primeira_linha=None`
            # (célula vazia depois continua aceita sem erro; célula com
            # QUALQUER valor não-vazio depois de uma primeira linha vazia
            # já é, por definição, DIFERENTE de `None` → conflito) — o
            # bug era só chamar a função condicionalmente. Mesmo padrão
            # incondicional já usado para `ap`/`bp` acima.
            _exigir_valor_repetido_igual(ws, ws_bruto, linha, 9,
                                         "nível d'água (m)", nivel_agua)
        camadas.append(Camada(nome=nome, espessura=espessura, tipo=tipo,
                              nspt=nspt, Cc=Cc, e0=e0, cv=cv, Es=Es))

    if not camadas:
        raise ValueError(f"Aba {ABA_PERFIL!r}: nenhuma camada encontrada.")

    return PerfilGeotecnico(camadas=camadas, nivel_agua=nivel_agua)


# --------------------------------------------------------------------------- #
#  Geração do modelo (.xlsx de exemplo + instruções)
# --------------------------------------------------------------------------- #
def gerar_modelo_importacao(caminho) -> None:
    """Escreve um `.xlsx` com as duas abas de dados (uma linha de exemplo
    preenchida em cada) e uma terceira aba "Instruções"."""
    livro = openpyxl.Workbook()

    ws_pilar = livro.active
    ws_pilar.title = ABA_PILAR
    ws_pilar.append(CABECALHO_PILAR)
    ws_pilar.append([0.30, 0.30, "G", 600.0, 15.0, 8.0])
    ws_pilar.append([None, None, "Q", 180.0, 6.0, 0.0])
    for cel in ws_pilar[1]:
        cel.font = Font(bold=True)

    ws_perfil = livro.create_sheet(ABA_PERFIL)
    ws_perfil.append(CABECALHO_PERFIL)
    ws_perfil.append(["Areia argilosa", "granular", 2.0, 12, None, None,
                      None, None, 1.5])
    ws_perfil.append(["Argila mole", "coesivo", 3.0, None, 0.45, 1.10,
                      0.02, None, None])
    for cel in ws_perfil[1]:
        cel.font = Font(bold=True)

    ws_instr = livro.create_sheet(ABA_INSTRUCOES)
    linhas = [
        "SAPATA-7 — modelo de importação de dados (escopo amplo).",
        "",
        f'Aba "{ABA_PILAR}":',
        "  Cabeçalho na linha 1, dados a partir da linha 2 — uma linha por caso de carga.",
        "  ap (m) / bp (m): dimensões do pilar em planta; só a PRIMEIRA linha de dados é lida (pode deixar em branco nas linhas seguintes).",
        "  caso: nome/rótulo do caso de carga. Para o caso importado aparecer AUTOMATICAMENTE nos campos da tela (Importar do Excel...),",
        "    o nome tem de ser exatamente 'G', 'Q' ou 'W' (maiúsculas) — a tela só tem esses três campos fixos (G sempre, Q e W opcionais).",
        "    Qualquer outro nome (ex. 'Permanente', 'Acidental') faz a importação de pilar/cargas FALHAR com um erro claro, em vez de entrar",
        "    parcialmente ou em silêncio — não use nomes livres nesta coluna se o destino é a tela do escopo amplo.",
        "  N (kN): carga axial — obrigatório em cada linha.",
        "  Mx, My (kN·m): momentos fletores — em branco vale 0.",
        "  Esta aba não tem colunas de Hx/Hy nem de tipo de ação: os casos importados entram como ação simples (tipo padrão: permanente).",
        "  Ajuste tipo/psi/forças horizontais na tela depois de importar, se o projeto precisar.",
        "",
        f'Aba "{ABA_PERFIL}":',
        "  Cabeçalho na linha 1, dados a partir da linha 2 — uma linha por camada, do TOPO para a BASE (a ordem das linhas é a ordem do perfil).",
        "  camada: nome da camada (texto livre).",
        f"  tipo: um de {TIPOS_ACEITOS} (valores do enum TipoSubstrato).",
        "  espessura (m): espessura da camada, > 0.",
        "  nspt: N_SPT médio da camada (opcional).",
        "  Cc, e0: índice de compressão e índice de vazios inicial — só fazem sentido para camada coesiva; deixe em branco nas demais.",
        "  cv (m²/ano): coeficiente de adensamento — só para camada coesiva; deixe em branco nas demais.",
        "    A unidade é m²/ano, a mesma que calc_core.sapata_isolada.geotecnia.Camada.cv espera: esta aba não converte unidade",
        "    nenhuma (regra de a3-interface.md: 'ui/ não calcula'), o valor digitado entra direto no núcleo.",
        "  Es (kPa): módulo de deformabilidade (opcional); se ausente e nspt estiver preenchido, o núcleo estima Es pela correlação com o SPT.",
        "  nível d'água (m): profundidade do N.A. a partir da superfície; só a PRIMEIRA linha de dados é lida.",
        "    Em branco na primeira linha = sem N.A. informado. Um valor DIFERENTE numa linha seguinte é um ERRO (a planilha",
        "    diz uma coisa, a importação usaria outra) — deixe em branco ou repita o mesmo valor.",
        "",
        "Domínio dos valores: espessura tem de ser > 0; nspt/Cc/e0/cv/Es/nível d'água têm de ser >= 0 quando preenchidos.",
        "Valores fora desse intervalo (ex.: espessura negativa por erro de digitação) são recusados na importação, não",
        "entram em silêncio no cálculo.",
        "",
        "Erros de preenchimento (célula obrigatória vazia, valor não numérico, valor lógico VERDADEIRO/FALSO onde se",
        "espera número, fórmula sem valor calculado, valor fora do domínio, tipo de substrato desconhecido, aba ausente)",
        "são reportados citando a aba, a linha e a coluna — nada é assumido em silêncio.",
    ]
    for i, texto in enumerate(linhas, start=1):
        ws_instr.cell(row=i, column=1, value=texto)
    ws_instr.column_dimensions["A"].width = 110

    for col, largura in zip("ABCDEF", (10, 10, 10, 10, 12, 12)):
        ws_pilar.column_dimensions[col].width = largura
    for col, largura in zip("ABCDEFGHI", (18, 12, 14, 8, 8, 8, 12, 10, 16)):
        ws_perfil.column_dimensions[col].width = largura

    livro.save(caminho)
